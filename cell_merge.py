from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet
from typing import List


def merge_same_value(from_source='schedule_t.xlsx', to_source='schedule.xlsx'):
    print(f"merge 시작 - {from_source} -> {to_source}")
    wb = load_workbook(from_source, read_only=False)
    merge_sheets(wb)
    wb.save(to_source)


def autofit_column(sheet, margin=10):
    for i, column_cells in enumerate(sheet.columns):
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + margin

def merge_sheets(wb):
    for s_name in wb.sheetnames:
        ws = wb[s_name]
        merge_sheet(ws)
        autofit_column(ws)


def merge_sheet(ws):
    key_columns = [1, 2]  # [대주제(Part), 중주제(Chapter)]
    for key_column in key_columns:
        merge_cell(ws, key_column)


def merge_cell(ws, key_column):
    max_row = ws.max_row
    prev_value = None
    start_row = 2
    for row, row_cells in enumerate(ws.iter_rows(min_col=key_column, min_row=start_row,
                                                 max_col=key_column, max_row=max_row),
                                    start_row):
        if prev_value != row_cells[0].value or row == max_row:
            if prev_value is not None:
                if row == max_row and prev_value == row_cells[0].value:
                    merge_segment(ws, start_row, row, key_column)
                else:
                    merge_segment(ws, start_row, row - 1, key_column)
                start_row = row
            prev_value = row_cells[0].value


def sum_lecture_time(l_times: List[str]):
    mm = 0
    ss = 0
    for l_time in l_times:
        t = l_time.split(':')
        mm += int(t[0])
        ss += int(t[1])
    mm += ss // 60
    ss = ss % 60
    hh = mm // 60
    mm = mm % 60
    return f'{hh:02d}:{mm:02d}:{ss:02d}' if hh != 0 else f'{mm:02d}:{ss:02d}'


def merge_segment(ws: Worksheet, start_row, end_row, key_column):
    ws.merge_cells(start_row=start_row, start_column=key_column, end_row=end_row,
                   end_column=key_column)

    current_cell = ws.cell(row=start_row, column=key_column)
    l_times = [l_time[0].value for l_time in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=4, max_col=4)]
    lec_name = f'{current_cell.value} ({sum_lecture_time(l_times)})'
    current_cell.value = lec_name
    current_cell.alignment = Alignment(horizontal='left', vertical='center')

if __name__ == '__main__':
    merge_same_value('lecture_t.xlsx', 'lecture.xlsx')
